# -*- coding: utf-8 -*-
"""
Created on Tue Oct 23 20:34:38 2018

@author: aniksinh
"""

import win32com.client
import imaplib
import re
import openpyxl as op
import xlrd
#import xlwt
#
#from xlwt import Workbook
#wb = Workbook()
#sheet1 = wb.add_sheet('Sheet 1') 

target ='D:\\Users\\aniksinh\\Desktop\\HR.xlsx'
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]

workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

inbox = outlook.Folders("anikesh.sinha@capgemini.com").Folders("HR Changes")
#inbox = outlook.GetDefaultFolder(1) # "6" refers to the index of a folder - in this case,
                                    # the inbox. You can change that number to reference
                                   # any other folder
accounts= win32com.client.Dispatch("Outlook.Application").Session.Accounts
messages = inbox.Items
a=len(messages)
print(type(messages))
print(messages)
message = messages.GetFirst()
while message:
    body_content = message.body
    b=list() 
    b=body_content.split("\r\n\r\n")
    b1=[i.split(':', 1)[1] for i in b]
    print(b1)
    message = messages.GetNext ()
            
#print(type(messages))
#for hr in messages :
#    message = messages.GetFirst()
#    print(type(message))
#body_content = message.body
#b=list() 
#b=body_content.split("\r\n\r\n")
#b1=[i.split(':', 1)[1] for i in b]
#print(b1)
#    message = messages.GetNext ()
#while message:
#    message = messages.GetNext()
#    print(b1)
#       print(b1)
#    for i in range(0,len(b1)) :
#        j=1
#        atr=b1[i]
#        print(atr)
#        ws.cell(row=j, column=i+1).value = atr
#        j=j+1
#        wb.save(target)
#        sheet1.write(1,atr,'b1[atr]')
#        wb.save('D:\\Users\\aniksinh\\Desktop\\HR.xlsx')
#         ws.cell(row=1, column=atr).value = 'b1[atr]'
#         wb.save(target)
    
#    d1=dict()
#    d1=b
#    
#    print (body_content)
#    print(c)
#    print(d1[0])
#    if b == d1 :
#        print(True)
#    i=1
#    for bd in len(b) :
#        print(b[i])
#        i+=2



#import win32com.client
#import xlwt
#outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
#inbox = outlook.Folders("IN, Clarity.BPOPoland").Folders("HR Current")
#book = xlwt.Workbook(encoding="utf-8")
#sheet = book.add_sheet("New Sheet")
#for folder in inbox :
#    fold = folder.Items
#    for messages in fold:
#        date = fold.ReceivedTime
#        sender = fold.Sender
#        sheet.write(1,0,date)
#        sheet.write(2,0,sender.Name)