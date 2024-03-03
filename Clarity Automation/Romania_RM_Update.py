# -*- coding: utf-8 -*-
"""
Created on Tue Nov 20 18:54:22 2018

@author: aniksinh
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import win32com.client
import openpyxl as op
import xlrd


# Define the target excel file
target ='D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
time.sleep(20)

No_rows1=sheet.nrows
for i in range(1,sheet.nrows):
  try:  
    #URL hit    
    driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
    time.sleep(5)
    
    #resourceid search   
    resid=sheet.cell_value(i ,0)
    driver.find_element_by_name("unique_name").clear() #clearing prev data
    driver.find_element_by_name("unique_name").send_keys(resid)    #sending res-id value to textbox
    driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
    
    #getting in resource
    driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
    time.sleep(5)
    
    #Properties Tab
        
    rmid=sheet.cell_value(i ,1)
    driver.find_element_by_name("manager_id_text").clear()
    driver.find_element_by_name("manager_id_text").send_keys(rmid)
    time.sleep(5)
    driver.find_element_by_name("manager_id_text").send_keys(Keys.ENTER)        #resource Manager id
    
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(5)                                                                #savebutton
    driver.find_element_by_link_text('Properties').click()
    driver.find_element_by_link_text('Resource').click()
    time.sleep(5)
#    driver.find_element_by_xpath("/html/body/div[1]/div[4]/div/div[3]/table/tbody/tr[3]/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(5)
    driver.find_element_by_xpath("//input[contains(@value, '3632')]").click()
    driver.find_element_by_xpath("//input[contains(@value, '3631')]").click()
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(5)
    driver.find_element_by_name("resourceIdFilter").send_keys(rmid)
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div/button[1]").click()
    time.sleep(5)
    driver.find_element_by_name("userId").click()    
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
    time.sleep(5)
    
    ws.cell(row=i+1, column=24).value = 'Resource Manager updated'
    wb.save(target)
  except:
      continue
  
wb.close()
driver.close()