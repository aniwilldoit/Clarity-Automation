# -*- coding: utf-8 -*-
"""
Created on Thu Oct 18 10:06:20 2018

@author: pdas7
"""

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl as op
import xlrd

# Define the target excel file
target ='C:\\Users\\pdas7\\Desktop\\clarity.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\Users\pdas7\AppData\Local\Continuum\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
time.sleep(5)

for i in range(1,sheet.nrows):
    #URL hit    
    driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
    time.sleep(5)
    
    #resourceid search   
    resid=sheet.cell_value(i ,0)
    driver.find_element_by_name("unique_name").clear() #clearing prev data
    driver.find_element_by_name("unique_name").send_keys(resid)    #sending res-id value to textbox
    driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
    
    #getting in resource
    driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
    time.sleep(3)
    
    prj=list()
    proj=sheet.cell_value(i,1)
    prj=proj.split(",")
    if len(prj) > 0:
        driver.find_element_by_link_text('Allocations').click()
        time.sleep(3)
        driver.execute_script("submitFormTarget('modal', 'page_projmgr.getResourceProjectObjectList','pma.selectResourceInvestments');")
  
        for j in range(0,len(prj)):
               try :     
                driver.find_element_by_name("unique_code").clear()
                driver.find_element_by_name("unique_code").send_keys(prj[j])
                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
                time.sleep(3)
                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/table/tbody/tr/td/div/table/thead/tr/th[1]/input").click()
               except NoSuchElementException as e:
                 continue
               
               if j < len(prj)-1:
                    driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[2]").click()
                    time.sleep(3)
               else:    
                    try:
                        driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()                
                        time.sleep(3)
                    except NoSuchElementException :
                        print(e)
        driver.execute_script("closeWindow('true');")
        time.sleep(2)
        ws.cell(row=i+1, column=3).value = 'Allocated'
        wb.save(target)

wb.close()
driver.close()