# -*- coding: utf-8 -*-
"""
Created on Fri Oct 19 16:18:21 2018

@author: pdas7
"""

from selenium import webdriver
#from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl as op
import xlrd


# Define the target excel file
target ='C:\\Users\\pdas7\\Desktop\\ResourceCreation1.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[2]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(2)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\Users\pdas7\AppData\Local\Continuum\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
time.sleep(20)

No_rows1=sheet.nrows
for i in range(1,sheet.nrows):
    try :
        #URL hit    
        driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
        time.sleep(5)
        
        #resourceid search   
        resid=sheet.cell_value(i ,10)
        print(resid)
        print(i)
        driver.find_element_by_name("unique_name").clear() #clearing prev data
        driver.find_element_by_name("unique_name").send_keys(resid)    #sending res-id value to textbox
        driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
        driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
        
        #getting in resource
        driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
        time.sleep(3)
        
        #Properties Tab
        doj=sheet.cell_value(i ,3)
        print(type(doj))
        print(doj)
        driver.find_element_by_name("date_of_hire").clear()
        driver.find_element_by_name("date_of_hire").send_keys(doj)
        driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
        time.sleep(2)
        ws.cell(row=i+1, column=23).value = 'GGID update success'
        wb.save(target)
    except NoSuchElementException :
        continue

wb.close()
driver.close()    