# -*- coding: utf-8 -*-
"""
Created on Sun Oct 15 12:55:52 2018

@author: aniksinh
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import win32com.client
import openpyxl as op
import xlrd


# Define the target excel file
target ='D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
time.sleep(20)

No_rows1=sheet.nrows
for i in range(1,sheet.nrows):
  try:  
    #URL hit    
    driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
    time.sleep(5)
    
    #resourceid search   
    resid=sheet.cell_value(i ,10).strip()
    driver.find_element_by_name("unique_name").clear() #clearing prev data
    driver.find_element_by_name("unique_name").send_keys(resid)    #sending res-id value to textbox
    driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
    
    #getting in resource
    driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
    time.sleep(5)
    
    #Properties Tab
    text4=sheet.cell_value(i ,0).strip()
    driver.find_element_by_name("prusertext4").clear()
    driver.find_element_by_name("prusertext4").send_keys(text4)     #text4
    
    doj=sheet.cell_value(i ,3).strip()
#    print(type(doj))
#    print(doj)
    driver.find_element_by_name("date_of_hire").clear()
    driver.find_element_by_name("date_of_hire").send_keys(doj)    #joining date
    
    text1=sheet.cell_value(i ,4).strip()
    driver.find_element_by_name("prusertext1").clear()
    driver.find_element_by_name("prusertext1").send_keys(text1)     #text1
    
    rmid=sheet.cell_value(i ,9).strip()
    driver.find_element_by_name("manager_id_text").clear()
    driver.find_element_by_name("manager_id_text").send_keys(rmid)
    time.sleep(5)
    driver.find_element_by_name("manager_id_text").send_keys(Keys.ENTER)        #resource Manager id
    
    grade=sheet.cell_value(i ,8).strip()
    driver.find_element_by_name("prcategory").clear()
    driver.find_element_by_name("prcategory").send_keys(grade)      #grade
    
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(5)                                                                #savebutton
    driver.find_element_by_link_text('Properties').click()
    driver.find_element_by_link_text('Resource').click()
    time.sleep(5)
    driver.find_element_by_xpath("/html/body/div[1]/div[4]/div/div[3]/table/tbody/tr[3]/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
    time.sleep(5)
    driver.find_element_by_xpath("//input[contains(@value, '3632')]").click()
    driver.find_element_by_xpath("//input[contains(@value, '3631')]").click()
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(5)
    driver.find_element_by_name("resourceIdFilter").send_keys(rmid)
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div/button[1]").click()
    time.sleep(5)
    driver.find_element_by_name("userId").click()    
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
    time.sleep(5)
    
    #Allocations Tab
    prj=list()
    proj=sheet.cell_value(i,18).strip()
    prj=proj.split(";")
    if prj[0] != '':
#        print(len(prj))
        driver.find_element_by_link_text('Allocations').click()
        time.sleep(7)
        driver.execute_script("submitFormTarget('modal', 'page_projmgr.getResourceProjectObjectList','pma.selectResourceInvestments');")
  
        for j in range(0,len(prj)):
               try :     
                driver.find_element_by_name("unique_code").clear()
                driver.find_element_by_name("unique_code").send_keys(prj[j])
                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
                time.sleep(5)
                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/table/tbody/tr/td/div/table/thead/tr/th[1]/input").click()
               except NoSuchElementException as e:
                continue
               
               if j < len(prj)-1:
                    driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[2]").click()
                    time.sleep(5)
               else:    
                    try:
                        driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()                
                        time.sleep(5)
                    except NoSuchElementException :
                        print(j)
        driver.execute_script("closeWindow('true');")
    ws.cell(row=i+1, column=23).value = 'Projects Allocated'
    wb.save(target)
        
#    driver.execute_script("closeWindow('true');")
    driver.find_element_by_link_text('Calendar').click()
    time.sleep(4)
    driver.find_element_by_name("baseCalendar").clear()             
    driver.find_element_by_name("baseCalendar").send_keys("Poland")
    time.sleep(3)
    driver.find_element_by_name("baseCalendar").send_keys(Keys.ENTER)
    driver.find_element_by_xpath("/html/body/div/div[4]/div[1]/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/table[1]/tbody/tr/td[3]/button").click()
    time.sleep(3)
    
    ws.cell(row=i+1, column=24).value = 'Resource created fully'
    wb.save(target)
  except:
      continue
  
wb.close()
driver.close()

#const=win32com.client.constants
#olMailItem = 0x0
#obj = win32com.client.Dispatch("Outlook.Application")
#newMail = obj.CreateItem(olMailItem)
#newMail.Subject = "Resource Creation sheet"
## newMail.Body = "I AM\nTHE BODY MESSAGE!"
#newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
#newMail.HTMLBody = "<HTML><BODY>Please find attached file below.</BODY></HTML>"
#newMail.To = "kalyan.singh@capgemini.com;jyothi.kunda@capgemini.com;ch-raja.adabala@capgemini.com;ashish.c.jha@capgemini.com"
##newMail.Cc= ""
#attachment1 = r"D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx"
#newMail.Attachments.Add(Source=attachment1)
#newMail.display(True)
#newMail.send()