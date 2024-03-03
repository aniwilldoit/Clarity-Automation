# -*- coding: utf-8 -*-
"""
Created on Sun Oct 14 09:19:41 2018

@author: pdas7
"""

from selenium import webdriver
import openpyxl as op
import time
import xlrd
#import pyautogui


# Define the target excel file
target ='C:\\Users\\pdas7\\Desktop\\ResourceCreation.xlsx'
 
# Open the workbook
workbook = xlrd.open_workbook(target)

#Open workbook for write
wk=op.Workbook()
wb = op.load_workbook(target) 

#Open active sheet for write
print(wb.sheetnames)
ws = wb.worksheets[0]
ws.cell(row=2, column=21).value = 'Edited'
wb.save(target)
wb.close()