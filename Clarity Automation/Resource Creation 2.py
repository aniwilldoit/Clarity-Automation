# -*- coding: utf-8 -*-
"""
Created on Sat Oct 13 16:42:16 2018

@author: aniksinh
"""

from selenium import webdriver
import openpyxl as op
import time
import xlrd
from selenium.common.exceptions import NoSuchElementException

# Define the target excel file
target ='D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx'
 
# Open the workbook
workbook = xlrd.open_workbook(target)

    ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]

    ##READ OPTIONS##
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
# Get single cell value (zero based)(testing read of data)
print(sheet.cell_value(1,19))
# Get total rows & total columns
print(sheet.nrows)
print(sheet.ncols)

#Driver path
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')

for i in range(1,sheet.nrows):
    #URL hit    
    driver.get('https://clarity.capgemini.com/niku/nu#action:odf.res_list_upgList')
    time.sleep(10)
    
    #username_search    
    resid=sheet.cell_value(i ,10)
    driver.find_element_by_name("kin").clear() #clearing prev data
    driver.find_element_by_name("kin").send_keys(resid)    #sending res-id value to textbox
    driver.find_element_by_xpath("//select[@name='status_1']/option[text()='All']").click()     #changing status to ALL from Active
    driver.find_element_by_xpath("//*[@id='page_13319155_collapseFilter']/div/button[1]").click()
    time.sleep(7)    
    #clicking Filter button
    try :
        #Fetching Error mesg by Created date Sort
        driver.find_element_by_link_text('Created Date').click()
        time.sleep(3)
        driver.find_element_by_link_text('Created Date').click()
        time.sleep(3)
        emsg=driver.find_element_by_class_name("ppm_read_only_value")
        print (emsg.text)   #console print of Error Message
        #Writing response to workbook   
        if emsg.text == 'User already exists in Clarity' :
            ws.cell(row=i+1, column=22).value = 'User already exists in Clarity'  #Checking for Error Message
            wb.save(target)
        else :
            ws.cell(row=i+1, column=22).value = 'Created successfully'            #Checking for No error Message
            wb.save(target)
    except NoSuchElementException :    
        ws.cell(row=i+1, column=22).value = 'Not Created'
        
wb.save(target) #saving the Excel file
time.sleep(2)
wb.close()  #closing used workbook
#driver.quit()   #killing Webdriver
driver.close()  #closing browser