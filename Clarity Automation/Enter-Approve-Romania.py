# -*- coding: utf-8 -*-
"""
Created on Tue Oct 23 11:58:45 2018

@author: pdas7
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import win32com.client
import openpyxl as op
import xlrd


# Define the target excel file
target ='D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:odf.customObject&odf_code=res_list_upg&odf_parent_id=5114737&id=5114737')
time.sleep(20)

No_rows1=sheet.nrows
for i in range(1,sheet.nrows):
  try:  
#    #URL hit    
#    driver.get('https://clarity.capgemini.com/niku/nu#action:odf.res_list_upgProperties&id=5114737&odf_view=res_list_upgProperties&odf_code=res_list_upg')
#    time.sleep(5)
#    
    resid=sheet.cell_value(i ,0)
    driver.find_element_by_name("fs_attr_ts_edit_ins_text_entry").send_keys(resid)
    time.sleep(2)
    driver.find_element_by_name("fs_attr_ts_edit_ins_text_entry").send_keys(Keys.ENTER)
    time.sleep(1)
    driver.find_element_by_name("fs_attr_ts_appr_ins_text_entry").send_keys(resid)
    time.sleep(2)
    driver.find_element_by_name("fs_attr_ts_appr_ins_text_entry").send_keys(Keys.ENTER)
#    driver.execute_script("optionSelectAll('page','fs_attr_addremove_gr');optionSelectAll('page','fs_attr_ts_edit_ins');optionSelectAll('page','fs_attr_ts_appr_ins');optionSelectAll('page','fs_attr_hrdbk_ins');optionSelectAll('page','fs_attr_edit_cal_ins');optionSelectAll('page','fs_attr_prj_edit_ins');optionSelectAll('page','fs_attr_mod_basel_in');optionSelectAll('page','fs_attr_ts_enter_obs');optionSelectAll('page','fs_attr_ts_appr_obs');optionSelectAll('page','fs_attr_prj_edit_obs');submitForm('page','odf.customObjectUpdateAndClose','id=5109575','odf_pk=5109575','odf_view=res_list_upgProperties','odf_code=res_list_upg','odf_error_action=odf.res_list_upgProperties');")
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(3)
    ws.cell(row=i+1, column=3).value = 'Enter-Approve provided'
    wb.save(target)
  except NoSuchElementException :
      continue
    
driver.close()
#const=win32com.client.constants
#olMailItem = 0x0
#obj = win32com.client.Dispatch("Outlook.Application")
#newMail = obj.CreateItem(olMailItem)
#newMail.Subject = "Resource Creation sheet"
## newMail.Body = "I AM\nTHE BODY MESSAGE!"
#newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
#newMail.HTMLBody = "<HTML><BODY>Please find attached file below.</BODY></HTML>"
#newMail.To = "ch-raja.adabala@capgemini.com"
##newMail.Cc= ""
#attachment1 = r"D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx"
#newMail.Attachments.Add(Source=attachment1)
#newMail.display(False)
#newMail.send()