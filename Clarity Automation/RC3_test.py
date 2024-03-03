# -*- coding: utf-8 -*-
"""
Created on Wed Oct 17 03:56:35 2018

@author: pdas7
"""

# -*- coding: utf-8 -*-
"""
Created on Sun Oct 15 12:55:52 2018

@author: pdas7
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl as op
import xlrd


# Define the target excel file
target ='C:\\Users\\pdas7\\Desktop\\ResourceCreation1.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\Users\pdas7\AppData\Local\Continuum\Anaconda3\chromedriver.exe')


No_rows1=sheet.nrows
for i in range(1,sheet.nrows):
    #URL hit    
    driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
    time.sleep(15)
    
    #resourceid search   
    resid=sheet.cell_value(i ,10)
    driver.find_element_by_name("unique_name").clear() #clearing prev data
    driver.find_element_by_name("unique_name").send_keys(resid)    #sending res-id value to textbox
    driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
    
    #getting in resource
    driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
    time.sleep(3)
    
    #Properties Tab
    text4=sheet.cell_value(i ,0)
    driver.find_element_by_name("prusertext4").clear()
    driver.find_element_by_name("prusertext4").send_keys(text4)     #text4
    
#    doj=sheet.cell_value(i ,3)
#    print(type(doj))
#    print(doj)
#    driver.find_element_by_name("date_of_hire").send_keys(doj)      #joining date
    
    text1=sheet.cell_value(i ,4)
    driver.find_element_by_name("prusertext1").clear()
    driver.find_element_by_name("prusertext1").send_keys(text1)     #text1
    
    rmid=sheet.cell_value(i ,9)
    driver.find_element_by_name("manager_id_text").clear()
    driver.find_element_by_name("manager_id_text").send_keys(rmid)
    time.sleep(2)
    driver.find_element_by_name("manager_id_text").send_keys(Keys.ENTER)        #resource Manager id
    
    grade=sheet.cell_value(i ,8)
    driver.find_element_by_name("prcategory").clear()
    driver.find_element_by_name("prcategory").send_keys(grade)      #grade
    
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()     #savebutton
    driver.find_element_by_link_text('Properties').click()
    driver.find_element_by_link_text('Resource').click()
    time.sleep(2)
    driver.find_element_by_xpath("/html/body/div[1]/div[4]/div/div[3]/table/tbody/tr[3]/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
    time.sleep(2)
    driver.find_element_by_xpath("//input[contains(@value, '3632')]").click()
    driver.find_element_by_xpath("//input[contains(@value, '3631')]").click()
    driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
    time.sleep(2)
    driver.find_element_by_name("resourceIdFilter").send_keys(rmid)
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div/button[1]").click()
    time.sleep(2)
    driver.find_element_by_name("userId").click()    
    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
    time.sleep(2)
    
    #Allocations Tab
    prj=list()
    proj=sheet.cell_value(i,18)
    prj=proj.split(",")
    if len(prj) > 0:
        driver.find_element_by_link_text('Allocations').click()
        time.sleep(2)
        driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[2]").click()
        time.sleep(1)    
        driver.find_element_by_xpath("//*[@id='page_pma.selectResourceInvestments_collapseFilter_action_img']").click()
        for j in range(0,len(prj)):
            driver.find_element_by_name("unique_code").clear()
            driver.find_element_by_name("unique_code").send_keys(prj[j])
            driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
#            driver.find_element_by_class_name("checkbox").click()
            time.sleep(3)            
            driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/table/tbody/tr/td/div/table/thead/tr/th[1]/input").click()
            if j == len(prj)-1:
                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[3]").click()
            else:    
                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[2]").click()
                time.sleep(3)
    else:
        driver.find_element_by_link_text('Calendar').click()
        time.sleep(2)
        driver.find_element_by_name("baseCalendar").clear()             
        driver.find_element_by_name("baseCalendar").send_keys("Poland")
        time.sleep(2)
        driver.find_element_by_name("baseCalendar").send_keys(Keys.ENTER)
        driver.find_element_by_xpath("/html/body/div/div[4]/div[1]/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/table[1]/tbody/tr/td[3]/button").click()
        time.sleep(2)
    
    ws.cell(row=i+1, column=23).value = 'Allocated'
    wb.save(target)
  
wb.close()
driver.close()