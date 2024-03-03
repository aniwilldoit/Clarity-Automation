# -*- coding: utf-8 -*-
"""
Created on Fri Oct 26 19:40:34 2018

@author: pdas7
"""

import win32com.client
from win32com.client import Dispatch, constants
from selenium import webdriver
#from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl as op
import xlrd


# Define the target excel file
target ='C:\\Users\\pdas7\\Desktop\\GGID.xlsx'
   
   ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[1]
 
# Open the workbook
workbook = xlrd.open_workbook(target)
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(1)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\Users\pdas7\AppData\Local\Continuum\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:mainnav.work&classCode=project')
time.sleep(15)
driver.find_element_by_name("unique_code").clear()
driver.find_element_by_name("unique_code").send_keys("PL-ITO-PKDBSYST-OS")
#driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
#Checkmark
#
#No_rows1=sheet.nrows
#for i in range(13,sheet.nrows):
#    try :
#        #URL hit    
#        driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
#        time.sleep(5)
#        
#        #resourceid search   
#        resid=sheet.cell_value(i ,0)
#        driver.find_element_by_name("unique_name").clear() #clearing prev data
#        driver.find_element_by_name("unique_name").send_keys(resid)    #sending res-id value to textbox
#        driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
#        driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
#        
#        #getting in resource
#        driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
#        time.sleep(25)
#        
#        #Properties Tab
#        text4=str(int(sheet.cell_value(i ,1)))
#        print (type (text4))
#        print (text4)
#        driver.find_element_by_name("prusertext4").clear()
#        driver.find_element_by_name("prusertext4").send_keys(text4)     #text4
#        driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
#        time.sleep(2)
#        ws.cell(row=i+1, column=3).value = 'GGID update success'
#        wb.save(target)
#    except NoSuchElementException :
#        continue
#
#wb.close()
#driver.close()  
#
#const=win32com.client.constants
#olMailItem = 0x0
#obj = win32com.client.Dispatch("Outlook.Application")
#newMail = obj.CreateItem(olMailItem)
#newMail.Subject = "Resource creation"
## newMail.Body = "I AM\nTHE BODY MESSAGE!"
#newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
#newMail.HTMLBody = "<HTML><BODY>Please find attached file. GGID Update is complete</BODY></HTML>"
#newMail.To = "kalyan.singh@capgemini.com;jyothi.kunda@capgemini.com;ch-raja.adabala@capgemini.com;ashish.c.jha@capgemini.com"
#newMail.Cc= "anil.valecha@capgemini.com"
#attachment1 = r"C:\\Users\\pdas7\\Desktop\\GGID.xlsx"
#newMail.Attachments.Add(Source=attachment1)
#newMail.display(True)
#newMail.send()
#  