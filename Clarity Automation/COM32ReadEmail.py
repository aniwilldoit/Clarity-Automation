# -*- coding: utf-8 -*-
"""
Created on Tue Oct 23 20:34:38 2018

@author: aniksinh
"""

import win32com.client
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl as op
import xlrd

target ='D:\\Users\\aniksinh\\Desktop\\HR.xlsx'
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]

workbook = xlrd.open_workbook(target)
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')
 
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)

# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
time.sleep(20)

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

inbox = outlook.Folders("anikesh.sinha@capgemini.com").Folders("HR Changes")
#inbox = outlook.GetDefaultFolder(1) # "6" refers to the index of a folder - in this case,
                                    # the inbox. You can change that number to reference
                                   # any other folder
accounts= win32com.client.Dispatch("Outlook.Application").Session.Accounts
messages = inbox.Items
#a=len(messages)
print(type(messages))
print(messages)
message = messages.GetFirst()
j=0
while message:
    body_content = message.body
    b=list() 
    b=body_content.split("\r\n\r\n")
    b1=[i.split(':', 1)[1] for i in b]
    print(b1)
    j=j+1
    for i in range(0,len(b1)) :
        atr=b1[i].strip()
        print(atr)
        ws.cell(row=j+1, column=i+1).value = atr
        wb.save(target)
    message = messages.GetNext()       


time.sleep(5)
for i in range(1,sheet.nrows):
  try:  
    #URL hit    
    driver.get('https://clarity.capgemini.com/niku/nu#action:projmgr.getResources&reload=true')
    time.sleep(5)
    
    text4=list()
    txt4=sheet.cell_value(i ,5)
    text4=txt4.split(";")
    print(text4)
    if text4[0] != '':
        for j in range(0,len(text4)):
             try :     
                driver.find_element_by_name("prusertext4").clear()
                driver.find_element_by_name("prusertext4").send_keys(text4[j])
                driver.find_element_by_xpath("//select[@name='is_active']/option[text()='Yes']").click()
                driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
                time.sleep(5)
                driver.find_element_by_xpath("//*[@id='projmgr.editResource']").click()
                time.sleep(5)
                
                #resourceid search 
                new_manager=list()
                resid=sheet.cell_value(i ,6)
                if resid[5:6] == '\\' :
                    new_manager=resid.split("\\")[1]
                    print(new_manager)
                    flag=1
                    driver.find_element_by_name("manager_id_text").clear() #clearing prev data
                    driver.find_element_by_name("manager_id_text").send_keys(new_manager)    #sending res-id value to textbox
                    time.sleep(5)
                    driver.find_element_by_name("manager_id_text").send_keys(Keys.ENTER)
                else:
                    driver.find_element_by_name("manager_id_text").clear() #clearing prev data
                    driver.find_element_by_name("manager_id_text").send_keys(resid)    #sending res-id value to textbox
                    time.sleep(5)
                    driver.find_element_by_name("manager_id_text").send_keys(Keys.ENTER)
                    flag=0
                
                text1=sheet.cell_value(i ,11)
                driver.find_element_by_name("prusertext1").clear()
                driver.find_element_by_name("prusertext1").send_keys(text1)
                
                driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
                time.sleep(5)
                
                driver.find_element_by_link_text('Properties').click()
                driver.find_element_by_link_text('Resource').click()
                time.sleep(5)
                driver.find_element_by_xpath("/html/body/div[1]/div[4]/div/div[3]/table/tbody/tr[3]/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
                time.sleep(5)
                driver.find_element_by_xpath("//input[contains(@value, '3632')]").click()
                driver.find_element_by_xpath("//input[contains(@value, '3631')]").click()
                driver.find_element_by_xpath("/html/body/div/div[5]/div/button[1]").click()
                time.sleep(5)
                
                prj=list()
                proj=sheet.cell_value(i,10)
                prj=proj.split(";")
                if prj[0] != '':
            #        print(len(prj))
                    driver.find_element_by_link_text('Allocations').click()
                    time.sleep(7)
                    driver.execute_script("submitFormTarget('modal', 'page_projmgr.getResourceProjectObjectList','pma.selectResourceInvestments');")
              
                    for j in range(0,len(prj)):
                           try :     
                            driver.find_element_by_name("unique_code").clear()
                            driver.find_element_by_name("unique_code").send_keys(prj[j])
                            driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div[2]/div/button[1]").click()
                            time.sleep(5)
                            driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/table/tbody/tr/td/div/table/thead/tr/th[1]/input").click()
                           except NoSuchElementException as e:
                            continue
                           
                           if j < len(prj)-1:
                                driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[2]").click()
                                time.sleep(5)
                           else:    
                                try:
                                    driver.find_element_by_xpath("/html/body/div/div[6]/div[2]/div[2]/div[2]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()                
                                    time.sleep(5)
                                except NoSuchElementException :
                                    print(j)
                    driver.execute_script("closeWindow('true');")
                ws.cell(row=i+1, column=20).value = 'Projects Allocated'
                wb.save(target)
                
                if flag == 0:
                    driver.find_element_by_name("resourceIdFilter").send_keys(resid)
                    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div/button[1]").click()
                    time.sleep(5)
                else:
                    driver.find_element_by_name("resourceIdFilter").send_keys(new_manager)
                    driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[1]/div/button[1]").click()
                    time.sleep(5)
                    
                driver.find_element_by_name("userId").click()    
                driver.find_element_by_xpath("/html/body/div/div[4]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td/div/div/table/tbody/tr[2]/td/form/div[2]/div/button[1]").click()
                time.sleep(5)
                wb.close()
                
             except NoSuchElementException as e:
                continue
            
  except:
    print ("1")