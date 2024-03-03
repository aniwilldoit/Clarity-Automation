#Resource_Creation
"""
Created on Tue Sep 18 15:42:18 2018

@author: aniksinh
"""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import xlrd
import openpyxl as op

# Define the target excel file
target ='D:\\Users\\aniksinh\\Desktop\\ResourceCreation.xlsx'
 
     ##WRITE OPTIONS##
#Open workbook for write
wk = op.Workbook()
wb = op.load_workbook(target) 
#Open active sheet for write
ws = wb.worksheets[0]

# Open the workbook
workbook = xlrd.open_workbook(target)
# Open the sheet (zero based)
sheet = workbook.sheet_by_index(0)
 
# Get single cell value (zero based)
print(sheet.cell_value(1 ,0))
 
# Get total rows
print(sheet.nrows)

# Get total columns
print(sheet.ncols)
driver = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\chromedriver.exe')
driver.get('https://clarity.capgemini.com/niku/nu#action:odf.fs_obj_usersCreate&odf_object_name=fs_obj_users&portlet_instance_id=11900162&caReturnActionParams=odf.fs_obj_usersList&page_11900162_collapse=false&superSecretTokenKey=superSecretTokenValue&filterGridCode=fs_obj_usersList&isErrorRequest=false&caReturnAction=odf.fs_obj_usersList&odf_code=fs_obj_users&gridCode=fs_obj_usersList&partition_code=NIKU.ROOT&relabel=true&componentId=odf&componentCode=odf&instanceCode=FILTER_11900162&validationViewCode=null')
for i in range(1,sheet.nrows):  
    driver.get('https://clarity.capgemini.com/niku/nu#action:odf.fs_obj_usersCreate&odf_object_name=fs_obj_users&portlet_instance_id=11900162&caReturnActionParams=odf.fs_obj_usersList&page_11900162_collapse=false&superSecretTokenKey=superSecretTokenValue&filterGridCode=fs_obj_usersList&isErrorRequest=false&caReturnAction=odf.fs_obj_usersList&odf_code=fs_obj_users&gridCode=fs_obj_usersList&partition_code=NIKU.ROOT&relabel=true&componentId=odf&componentCode=odf&instanceCode=FILTER_11900162&validationViewCode=null')
    time.sleep(10)
    #UserName_resource    
    logid=sheet.cell_value(i ,5)
    print(logid)
    #driver.find_element_by_name("fs_attr_user_name").send_keys(logid)
    #manager_resouce_id
    resid=sheet.cell_value(i ,9)
    print(resid)
    #driver.find_element_by_name("manager_text").send_keys(resid)
#    time.sleep(5)
    #P&L OBS
    PLOBS=sheet.cell_value(i ,15).strip()
    print(PLOBS)
    #driver.find_element_by_name("pnl_obs_text").send_keys(PLOBS)
#    time.sleep(3)
    #RightShore OBS
    ROBS=sheet.cell_value(i ,16).strip()
    print(ROBS)
    #driver.find_element_by_name("rightshore_obs_text").send_keys(ROBS)
#    time.sleep(3)
    #Discipline OBS
    DOBS=sheet.cell_value(i ,17).strip()
    print(DOBS)
    #driver.find_element_by_name("discipline_obs_text").send_keys(DOBS)
#    time.sleep(3)
     #username   
    driver.find_element_by_name("fs_attr_user_name").send_keys(logid)
    #    #manager_text
    driver.find_element_by_name("manager_text").send_keys(resid)
    time.sleep(2)
    driver.find_element_by_name("manager_text").send_keys(Keys.ENTER)
    #    #company Capgemini
    driver.find_element_by_name("company").send_keys(Keys.ENTER)
    time.sleep(2)
    driver.find_element_by_xpath("//select[@name='company']/option[text()='Capgemini']").click()
    time.sleep(2)
    #    #time_zone
    driver.find_element_by_xpath("//select[@name='time_zone']/option[text()='(GMT+01:00) Sarajevo, Skopje, Sofija, Vilnius, Warsaw, Zagreb']").click()
    #    #fs_attr_locale
    time.sleep(2)
    driver.find_element_by_xpath("//select[@name='fs_attr_locale']/option[text()='Polish(Poland)']").click()
        #z_fs_attr_language
    time.sleep(2)
    driver.find_element_by_xpath("//select[@name='z_fs_attr_language']/option[text()='English']").click()
    time.sleep(2)
    driver.find_element_by_name("discipline_obs_text").send_keys(DOBS)
    time.sleep(4)
    driver.find_element_by_name("discipline_obs_text").send_keys(Keys.ENTER)
    time.sleep(3)
    driver.find_element_by_name("pnl_obs_text").send_keys(PLOBS)
    time.sleep(4)
    driver.find_element_by_name("pnl_obs_text").send_keys(Keys.ENTER)
    time.sleep(3)
    driver.find_element_by_name("rightshore_obs_text").send_keys(ROBS)
    time.sleep(4)
    driver.find_element_by_name("rightshore_obs_text").send_keys(Keys.ENTER)
    time.sleep(3)
    driver.find_element_by_name("timesheet_obs_text").send_keys("Default Process")
    time.sleep(2)
    driver.find_element_by_name("timesheet_obs_text").send_keys(Keys.ENTER)
    time.sleep(2)
    driver.find_element_by_name("add_group_user_text_entry").send_keys("Time_Recorder")
    time.sleep(2)
    driver.find_element_by_name("add_group_user_text_entry").send_keys(Keys.ENTER)
    time.sleep(4)
    driver.execute_script("optionSelectAll('page','add_group_user');optionSelectAll('page','edit_ts');optionSelectAll('page','timesheet_approve');optionSelectAll('page','hard_book');optionSelectAll('page','edit_calender');optionSelectAll('page','modify_baseline');optionSelectAll('page','project_edit');optionSelectAll('page','obs_timesheet_enter');optionSelectAll('page','obs_project_edit');optionSelectAll('page','obs_timesheet_approv');submitForm('page','odf.customObjectInsertAndClose','odf_view=fs_obj_usersCreate','odf_code=fs_obj_users','odf_error_action=odf.fs_obj_usersCreate');") 
    time.sleep(2)
    ws.cell(row=i+1, column=21).value = 'Created'
    wb.save(target)
    time.sleep(2)


time.sleep(2)
driver.close()